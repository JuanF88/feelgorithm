#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Genera src/data/matriz.js desde 'assets/Noticias - rutas.xlsx'.
Por caso (ID) y ruta emocional: distorsiones, decisiones→consecuencia, ejercicio, consejo.
Uso:  python3 _generar.py     (requiere: python3 -m pip install --user openpyxl)
"""
import openpyxl, json, os
ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
XLSX = os.path.join(ROOT, 'assets', 'Noticias - rutas.xlsx')
OUT = os.path.join(ROOT, 'src', 'data', 'matriz.js')

wb = openpyxl.load_workbook(XLSX, data_only=True); ws = wb.worksheets[0]
grid = [[c.value for c in row] for row in ws.iter_rows()]
for mr in ws.merged_cells.ranges:
    v = grid[mr.min_row-1][mr.min_col-1]
    for r in range(mr.min_row-1, mr.max_row):
        for c in range(mr.min_col-1, mr.max_col):
            grid[r][c] = v

def nid(v):
    if v is None: return None
    if isinstance(v, float) and v.is_integer(): return str(v)
    return str(v).strip()
def norm(v): return ' '.join(str(v).split()) if v is not None else ''

M = {}
for i in range(1, len(grid)):
    idv, emo = nid(grid[i][0]), grid[i][4]
    if not idv or not emo: continue
    node = M.setdefault(idv, {}).setdefault(str(emo).strip(),
        {'distorsiones': [], 'decisiones': {}, 'ejercicio': '', 'consejo': ''})
    if grid[i][5]:
        d = norm(grid[i][5])
        if d not in node['distorsiones']: node['distorsiones'].append(d)
    if grid[i][9] and grid[i][10]:
        node['decisiones'][str(grid[i][9]).strip()] = norm(grid[i][10])
    if grid[i][11] and not node['ejercicio']: node['ejercicio'] = norm(grid[i][11])
    if grid[i][12] and not node['consejo']: node['consejo'] = norm(grid[i][12])

hdr = ("// GENERADO desde 'assets/Noticias - rutas.xlsx' (regenerar con\n"
       "// _source/distorsiones/_generar.py). No editar a mano.\n"
       "// Por caso (ID) y emoción: distorsiones, decisiones→consecuencia, ejercicio, consejo.\n\n")
os.makedirs(os.path.dirname(OUT), exist_ok=True)
open(OUT, 'w').write(hdr + 'export const MATRIZ = ' + json.dumps(M, ensure_ascii=False, indent=2) + ';\n')
print('OK ->', OUT)
